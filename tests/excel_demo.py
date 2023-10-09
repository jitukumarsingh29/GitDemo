import openpyxl

book = openpyxl.load_workbook("PythonDDD.xlsx")
sheet = book.active
dict_test = {}
for i in range(1, sheet.max_row+1): # to get row
    if sheet.cell(row=i, column=1).value == "TestCase2":
        for j in range(2, sheet.max_column+1):
            dict_test[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(dict_test)