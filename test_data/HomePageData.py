import openpyxl


class HomePageData:

    test_home_page_data = [{"name": "Jitu", "email": "Kumar@test.com", "password": "Singh_pwd", "gender": "Male"}, {"name": "Aditi", "email": "aditi@test.com", "password": "Singh_pwd", "gender": "Female"}]

    @staticmethod
    def get_test_excel_data(test_case_name):
        dict_test = {}
        book = openpyxl.load_workbook("PythonDDD.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get row
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict_test[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [dict_test]  # we need to send data as list so wrapping dict object inside list object
